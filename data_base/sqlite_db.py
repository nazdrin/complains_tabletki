import sqlite3 as sq
import openpyxl
import json
#привет
#kuku

#dhfdjjf
def sql_start(): #создание базы данных и таблиц
    global base
    global cur
    base = sq.connect('data.db')
    cur = base.cursor()

    base.execute('CREATE TABLE IF NOT EXISTS {} (code_store P'
                 'RIMARY KEY, address_store, code_chain, name_chain, manager)'.format('pharma'))
    base.commit()
    base.execute('CREATE TABLE IF NOT EXISTS {} (number_request PRIMARY KEY, dat'
                 'e_request, date, sourse,  type_request, type_complaint, client, email, phone, number_order'
                 ', code_store, text_request, text_answer, status)'
                 .format('requests'))
    base.commit()
    #load_xls()

def sql_insert_request(number, date_request, date, sourse, type_appeal, type_complains, client, email, phone, number_order, code_store, text_appeal, text_answer,status):
    cur.execute('INSERT INTO requests VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                (number, date_request, date, sourse, type_appeal, type_complains, client, email, phone, number_order, code_store, text_appeal, text_answer,status))
    base.commit()
def sql_select_pharma(code_store):
    report = cur.execute('SELECT address_store, code_chain, name_chain,'
                         ' manager FROM pharma WHERE code_store == ?', (code_store,)).fetchall()
    report = report[0]
    return report
def sql_select_requests(status):
    no_finish = cur.execute('SELECT number_request, client, phone, email, code_store'
                         ' manager FROM request WHERE status == ?', (status,)).fetchall()
    return no_finish
def load_xls ():

    wb = openpyxl.load_workbook('Pharma.xlsx')  # Заполняем словарь pharmacies_chain из єксель при перезапуске бота
    sheet = wb.active
    rows = sheet.max_row

    for i in range(1, rows + 1):
        a = sheet.cell(row=i, column=1)
        b = sheet.cell(row=i, column=2)
        c = sheet.cell(row=i, column=3)
        d = sheet.cell(row=i, column=4)
        e = sheet.cell(row=i, column=5)
        code = a.value
        address = b.value
        code_chain = c.value
        name_chain = d.value
        manager = e.value
        cur.execute('INSERT INTO pharma VALUES (?, ?, ?, ?, ?)', (code, address, code_chain, name_chain, manager))
        base.commit()
def number_json_load ():# изменение номера запроса
    with open('number.txt') as json_file:
        number_request = json.load(json_file)

    return number_request


def number_json_save(number):  # изменение номера запроса
    with open('number.txt', 'w') as outfile:
        json.dump(number, outfile)
